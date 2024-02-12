import os

import customtkinter as ctk
import sys
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from Database.database_manager import DatabaseManager
from Database.quoting_db import QuoteDatabase


# import other GUIs


class GUIManager:
    def __init__(self, app, db_manager, user_authentication):
        self.app = app
        self.db_manager = db_manager
        self.user_authentication = user_authentication
        self.quote_db = QuoteDatabase

        # create frames
        self.login_frame = None

        self.create_login_ui()

    # All the hiding and closing frame methods
    def hide_create_user_ui(self):  # hide the create a user frame
        if self.create_user_frame:
            self.create_user_frame.destroy()
        self.create_login_ui()

    def hide_login_ui(self):  # Hiding the login frame
        if self.login_frame:
            self.login_frame.destroy()
        self.app.deiconify()

    def on_close(self):  # Closing the root from any window
        quit()

    def toggle_password(self):  # toggle button to make the password hidden or unhidden
        self.show_password = not self.show_password
        if self.show_password:
            self.password_entry.config(show="")
            # custom button here
        else:
            self.password_entry.config(show="*")
            # custom button

    # Login page
    def create_login_ui(self):
        # Login GUI
        self.app.withdraw()
        self.login_frame = ctk.CTkToplevel(self.app)
        self.login_frame.title("Login")
        self.login_frame.geometry("275x175")
        self.login_frame.protocol("WM_DELETE_WINDOW", self.on_close)

        # Entry box for the username
        self.username_entry = ctk.CTkEntry(self.login_frame, placeholder_text="Username")
        self.username_entry.pack(pady=8, padx=4)

        # Entry box for the password
        self.password = ctk.CTkEntry(self.login_frame, placeholder_text="Password", show="*")
        self.password.pack(pady=8, padx=4)

        # login button to confirm login in
        self.login_button = ctk.CTkButton(self.login_frame, text="Login",
                                          command=lambda: self.handle_login(self.username_entry.get(),
                                                                            self.password.get()))
        self.login_button.pack(pady=8, padx=4)
        # create users tab to switch over to the create frame
        self.create_user = ctk.CTkLabel(self.login_frame, text="Create Account", cursor="hand2")
        self.create_user.pack(pady=8, padx=4)
        self.create_user.bind("<Button-1>", lambda event: self.show_create_user(event) and self.hide_login_ui())

        self.create_user.configure(font=("Helvetica", 10, "underline"))
        # frame for the creation user
        self.create_user_frame = None

        # Bind the login to the enter key
        self.password.bind("<Return>",
                           lambda event: self.handle_login(self.username_entry.get(), self.password.get()))

    # Login
    def handle_login(self, username, password):
        if self.user_authentication.login(username, password):
            # Login sucessful
            messagebox.showinfo("Success", "Welcome to MVPS quoting")
            self.main_menu()
        else:
            # Login Fail
            self.password.delete(0, tk.END)
            messagebox.showerror("Login Failed", "Invalid username or password")

    # Creating a new user window

    def show_create_user(self, event):
        # GUI for the creating a new user page
        if self.create_user_frame is None or not self.create_user_frame.winfo_exists():
            self.hide_login_ui()
            self.app.withdraw()
            self.create_user_frame = ctk.CTkToplevel(self.create_user_frame)
            self.create_user_frame.protocol("WM_DELETE_WINDOW", self.on_close)
            self.create_user_frame.title("Create User")
            self.create_user_frame.geometry("275x215")

            # Entry box for new username
            self.create_user_entry = ctk.CTkEntry(self.create_user_frame, placeholder_text="Username")
            self.create_user_entry.pack(pady=8, padx=4)

            # Entry box for new user password
            self.password_entry = ctk.CTkEntry(self.create_user_frame, placeholder_text="Password", show="*")
            self.password_entry.pack(pady=8, padx=4)

            # Confirm password entry
            self.password_confirm = ctk.CTkEntry(self.create_user_frame, placeholder_text="Confirm Password",
                                                 show="*")
            self.password_confirm.pack(pady=8, padx=4)

            # Submit button to enter the new user
            self.create_button = ctk.CTkButton(self.create_user_frame, text="Create",
                                               command=lambda: self.create(self.create_user_entry.get(),
                                                                           self.password_entry.get(),
                                                                           self.password_confirm.get()))
            self.create_button.pack(pady=8, padx=4)

            # go back to login page label and link
            self.back_to_login = ctk.CTkLabel(self.create_user_frame, text="Back to login", cursor="hand2")
            self.back_to_login.pack(pady=8, padx=4)
            self.back_to_login.bind("<Button-1>", lambda event: self.hide_create_user_ui() and self.login_frame())

            self.back_to_login.configure(font=("Helvetica", 10, "underline"))
            # hides login frame when using create user page

    # Create
    def create(self, username, password, confirm):
        if self.user_authentication.create(username, password, confirm):
            messagebox.showinfo("Success", "account created successfully")
            self.hide_create_user_ui()
        else:
            messagebox.showerror("Error", "failed to create a user")

    # Main Menu of the application
    def main_menu(self):
        self.main_frame = ctk.CTkToplevel(self.app)
        self.main_frame.geometry("750x500")
        self.main_frame.title("Quoting")
        self.main_frame.protocol("WM_DELETE_WINDOW", self.on_close)

        # Create a new quote button
        self.create_quote = ctk.CTkButton(self.main_frame, command=self.on_click, text="Create Quote")
        self.create_quote.pack(pady=8, padx=4)

    def open_and_save(self):
        try:
            wb = load_workbook(r"C:\Users\adam\PycharmProjects\quoting\quotes\blank quote.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active

        quote_id = QuoteDatabase.get_quote_id(self)

        ws['h4'] = quote_id
        quote_id = ws['h4'].value

        self.quote_db.add_quote_id(self, quote_id)

        wb.save(r'C:\Users\adam\PycharmProjects\quoting\quotes\new quote.xlsx')

    def on_click(self):
        self.open_and_save()

        os.startfile(r"C:\Users\adam\PycharmProjects\quoting\quotes\new quote.xlsx")

